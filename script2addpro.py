import hashlib
import json
import time
import traceback
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
from pathlib import Path

import random

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from PIL import Image
import requests
from openpyxl.drawing.image import Image as OpenpyxlImage
from io import BytesIO

import sqlite3
# ===================== 必填配置 =====================

# 填写拼多多开放平台应用信息。
# 你之前公开过旧密钥，请使用后台重新生成的新密钥。
CLIENT_ID = "2f1a1d317468465e88f1a23a06dbf53e"
CLIENT_SECRET = "da868ab1476b301ad35351c2d9c65ec82a2be547"

# # 填写多多进宝推广位信息。
# PID = "44575313_316926926"
# UID = "44575313"

# 填写多多进宝推广位信息【多PID轮询池子】
# 把后台新建的所有完整PID全部写进列表，越多越不容易限流
PID_POOL = [
   # "44575313_316926926",   # 你原来在用的PID
    "44575313_316913643",        # 替换成你新建第1个PID
    "44575313_316942052",        # 替换成你新建第2个PID
    "44575313_316942048",   ]     # 替换成你新建第3个PID

UID = "44575313"
# 每次调用接口自动随机挑选一个PID使用
PID = random.choice(PID_POOL)

# 搜索关键词，可以填写多个。
KEYWORD_LIST = [
    "女装韩式风格 20-40元",
   # "微胖女生韩系穿搭平价",
]

PAGE_SIZE = 15
MAX_PAGE = 10

# Excel 保存位置。
SAVE_PATH = Path(r"D:\Excel\拼多多商品汇总8.xlsx")

# 拼多多开放平台网关。
GATEWAY = "https://gw-api.pinduoduo.com/api/router"


# ===================== 网络配置 =====================

# 忽略 Windows、PyCharm 或代理软件遗留的无效代理环境变量。
# 这正是解决你当前 ProxyError 的关键配置。
HTTP_SESSION = requests.Session()
HTTP_SESSION.trust_env = False

HTTP_SESSION.headers.update({
    "User-Agent": "Mozilla/5.0 PDD-OpenAPI-Client/1.0",
    "Accept": "application/json",
})

# ===================== 去重SQLite初始化 =====================
# 连接数据库
duplicate_conn = sqlite3.connect("spider_duplicate.db")
duplicate_cur = duplicate_conn.cursor()
# 创建已爬商品ID表，goods_id唯一约束
duplicate_cur.execute('''
CREATE TABLE IF NOT EXISTS crawled_goods (
    goods_id TEXT PRIMARY KEY
)
''')
duplicate_conn.commit()

def is_goods_exist(goods_id):
    """查询该商品是否已经抓取过"""
    duplicate_cur.execute("SELECT 1 FROM crawled_goods WHERE goods_id = ?", (str(goods_id),))
    return duplicate_cur.fetchone() is not None

def mark_goods_crawled(goods_id):
    """标记商品已抓取，重复则自动忽略"""
    duplicate_cur.execute("INSERT OR IGNORE INTO crawled_goods(goods_id) VALUES (?)", (str(goods_id),))
    duplicate_conn.commit()

# ===================== 工具函数 =====================

def validate_config():
    """运行前检查必要配置。"""

    if not CLIENT_ID or CLIENT_ID == "填写你的Client ID":
        raise RuntimeError("请先填写 CLIENT_ID")

    if not CLIENT_SECRET or CLIENT_SECRET == "填写你重置后的Client Secret":
        raise RuntimeError("请先填写 CLIENT_SECRET")

    if not PID:
        raise RuntimeError("请先填写 PID")

    if not UID:
        raise RuntimeError("请先填写 UID")


def normalize_sign_value(value):
    """
    将参数值转换为签名所需字符串。

    bool 必须处理为小写字符串，其他值直接使用 str。
    """
    if isinstance(value, bool):
        return "true" if value else "false"

    return str(value)


def create_sign(params):
    """
    拼多多开放平台签名规则：

    CLIENT_SECRET
    + 按参数名升序拼接 key 和 value
    + CLIENT_SECRET
    最后计算大写 MD5。
    """

    pieces = [CLIENT_SECRET]

    for key in sorted(params.keys()):
        value = params[key]

        if value is None:
            continue

        pieces.append(key)
        pieces.append(normalize_sign_value(value))

    pieces.append(CLIENT_SECRET)

    sign_source = "".join(pieces)

    return hashlib.md5(
        sign_source.encode("utf-8")
    ).hexdigest().upper()


def call_pdd_api(api_type, business_params):
    """调用拼多多开放平台接口。"""

    params = {
        "client_id": CLIENT_ID,
        "timestamp": int(time.time()),
        "type": api_type,
        "data_type": "JSON",
    }

    params.update(business_params)
    params["sign"] = create_sign(params)

    print(f"正在调用接口：{api_type}")

    try:
        response = HTTP_SESSION.post(
            GATEWAY,
            data=params,
            timeout=(10, 30),
        )

        response.raise_for_status()

    except requests.exceptions.ProxyError as exc:
        raise RuntimeError(
            "代理连接失败。当前代码已经设置 trust_env=False；"
            "请检查是否有代理软件强制接管系统流量。"
        ) from exc

    except requests.exceptions.SSLError as exc:
        raise RuntimeError(
            "HTTPS 握手失败。请检查代理软件、VPN、防火墙或杀毒软件的 HTTPS 扫描功能。"
        ) from exc

    except requests.exceptions.ConnectTimeout as exc:
        raise RuntimeError(
            "连接拼多多网关超时，请检查网络是否能访问 "
            "https://gw-api.pinduoduo.com"
        ) from exc

    except requests.exceptions.ReadTimeout as exc:
        raise RuntimeError("拼多多接口响应超时，请稍后重试") from exc

    except requests.exceptions.RequestException as exc:
        raise RuntimeError(f"网络请求失败：{exc}") from exc

    try:
        result = response.json()
    except ValueError as exc:
        raise RuntimeError(
            "拼多多接口没有返回 JSON。\n"
            f"HTTP 状态码：{response.status_code}\n"
            f"返回内容：{response.text[:1000]}"
        ) from exc

    if "error_response" in result:
        error = result["error_response"]

        error_code = error.get("error_code", "")
        sub_code = error.get("sub_code", "")
        error_msg = error.get("error_msg", "")
        sub_msg = error.get("sub_msg", "")
        request_id = error.get("request_id", "")

        raise RuntimeError(
            "\n拼多多接口返回错误：\n"
            f"error_code：{error_code}\n"
            f"sub_code：{sub_code}\n"
            f"error_msg：{error_msg}\n"
            f"sub_msg：{sub_msg}\n"
            f"request_id：{request_id}"
        )

    return result


# ===================== 商品搜索 =====================

def pdd_goods_search(keyword, page, page_size):
    business_params = {
        "keyword": keyword,
        "page": page,
        "page_size": page_size,
        "pid": PID,
    }

    # 第一次测试时设为 False，只验证 PID。
    # PID 单独成功后，再改成 True 验证 UID。
    ENABLE_CUSTOM_PARAMETERS = False

    if ENABLE_CUSTOM_PARAMETERS:
        if not UID:
            raise RuntimeError(
                "ENABLE_CUSTOM_PARAMETERS=True 时必须填写 UID"
            )

        business_params["custom_parameters"] = json.dumps(
            {"uid": str(UID)},
            ensure_ascii=False,
            separators=(",", ":"),
        )

    print(
        f"正在搜索：keyword={keyword!r}, "
        f"page={page}, page_size={page_size}, "
        f"pid={PID}, "
        f"custom_parameters={'启用' if ENABLE_CUSTOM_PARAMETERS else '停用'}"
    )

    return call_pdd_api(
        "pdd.ddk.goods.search",
        business_params,
    )


def extract_goods_list(result):
    """从接口响应中提取商品列表。"""

    search_response = result.get("goods_search_response")

    if not isinstance(search_response, dict):
        raise RuntimeError(
            "接口响应中没有 goods_search_response：\n"
            + json.dumps(result, ensure_ascii=False, indent=2)[:2000]
        )

    goods_list = search_response.get("goods_list", [])

    if goods_list is None:
        return []

    if not isinstance(goods_list, list):
        raise RuntimeError("goods_list 的数据格式不是列表")

    return goods_list


# ===================== 数据处理 =====================

def cent_to_yuan(value):
    """将分转换为元。"""

    try:
        return int(value or 0) / 100
    except (TypeError, ValueError):
        return 0


def build_goods_row(keyword, page, item):
    """将单个商品转换为 Excel 行。"""

    price_cent = int(item.get("min_group_price") or 0)
    coupon_cent = int(item.get("coupon_discount") or 0)
    final_cent = max(price_cent - coupon_cent, 0)

    # 提取官方主图链接填入I列
    main_img_url = item.get("goods_image_url", "")

    return [
        keyword,
        page,
        str(item.get("goods_id") or ""),
        item.get("goods_sign") or "",
        item.get("goods_name") or "",
        price_cent / 100,
        coupon_cent / 100,
        final_cent / 100,
        "",
        item.get("mall_name") or "",
        item.get("sales_tip") or "",
        item.get("promotion_rate") or 0,
    ]


def format_worksheet(sheet):
    """设置 Excel 表格的基本样式。"""

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="4472C4",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    widths = {
        "A": 15,
        "B": 8,
        "C": 22,
        "D": 45,
        "E": 60,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 60,
        "J": 30,
        "K": 15,
        "L": 15,
    }

    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for row_number in range(2, sheet.max_row + 1):
        sheet.cell(row_number, 6).number_format = "0.00"
        sheet.cell(row_number, 7).number_format = "0.00"
        sheet.cell(row_number, 8).number_format = "0.00"

    # 统一设置行高，适配图片
    for r in range(2, sheet.max_row + 1):
        sheet.row_dimensions[r].height = 80


# ===================== 主程序 =====================

def main():
    validate_config()

    print("=" * 60)
    print("程序开始运行")
    print(f"关键词数量：{len(KEYWORD_LIST)}")
    print(f"每页数量：{PAGE_SIZE}")
    print(f"最大页数：{MAX_PAGE}")
    print(f"保存位置：{SAVE_PATH}")
    print("=" * 60)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "商品汇总"

    sheet.append([
        "搜索词",
        "页码",
        "商品ID",
        "goods_sign",
        "商品标题",
        "拼团价(元)",
        "优惠券(元)",
        "券后价(元)",
        "主图链接",
        "店铺名称",
        "销量提示",
        "佣金比例",
    ])

    total = 0

    for keyword in KEYWORD_LIST:
        print(f"\n开始抓取关键词：{keyword}")

        for page in range(1, MAX_PAGE + 1):
            # 每页请求前随机更换PID，分摊接口调用量
            PID = random.choice(PID_POOL)
            print(f"当前本轮使用PID：{PID}")
            try:
                result = pdd_goods_search(
                    keyword=keyword,
                    page=page,
                    page_size=PAGE_SIZE,
                )
            except RuntimeError as err:
                print(f"\n❌ {keyword} 第{page}页调用接口失败，跳过该页：{err}")
                workbook.save(SAVE_PATH)
                #time.sleep(1)
                # 随机等待 1 ~ 3 秒
                time.sleep(random.uniform(1, 3))
                continue

            goods_list = extract_goods_list(result)

            if not goods_list:
                print(f"{keyword} 第 {page} 页没有商品，停止翻页")
                break

            page_count = 0

            for item in goods_list:
                price_cent = int(item.get("min_group_price") or 0)
                coupon_cent = int(item.get("coupon_discount") or 0)
                final_price = (price_cent - coupon_cent) / 100

                # 核心价格筛选：券后价必须 10 ≤ 价格 ≤ 20
                if not (10 <= final_price <= 20):
                    continue

                # 关键词二次兜底过滤（标题包含韩系/韩国/微胖/大码）
                title = item.get("goods_name", "").lower()
                filter_words = {"韩系", "韩国", "微胖", "大码", "胖mm", "胖妹妹"}
                if not any(word in title for word in filter_words):
                    continue

                # ========== 新增去重判断 ==========
                gid = item.get("goods_id")
                if is_goods_exist(gid):
                    print(f"商品{gid}已存在，跳过本条")
                    continue
                # 标记为已抓取
                mark_goods_crawled(gid)
                # =================================

                row_data = build_goods_row(keyword, page, item)
                # 先写入一行空数据
                sheet.append(row_data)
                current_row = sheet.max_row
                img_url = item.get("goods_image_url")
               # imgs_url=item.get("goods_images_url")
               # multi_imgs = item.get("goods_gallery_urls", "").split(",")
                gallery_raw = item.get("goods_gallery_urls", "")
                multi_imgs = [link.strip() for link in gallery_raw.split(",") if link.strip()]
                if img_url:
                    try:
                        # 加上浏览器请求头，绕过拼多多图片防盗链
                        headers = {
                            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                            "Referer": "https://mobile.yangkeduo.com/"
                        }

                        # 单张图片请求随机停顿0.5~2秒，模拟人工访问
                        time.sleep(random.uniform(0.5, 2.0))

                        resp_img = requests.get(img_url, headers=headers, timeout=8)
                        resp_img.raise_for_status()
                        img_stream = BytesIO(resp_img.content)
                        img = OpenpyxlImage(img_stream)
                        img.width = 160
                        img.height = 160
                        sheet.add_image(img, f"I{current_row}")
                    except Exception as e:
                        print(f"图片拉取失败 {img_url} | 错误：{str(e)}")
                total += 1
                page_count += 1

            print(
                f"{keyword} 第 {page} 页完成，"
                f"本页 {page_count} 条，累计 {total} 条"
            )
            # 每抓取一页立刻保存文件，防止报错丢数据
            # workbook.save(SAVE_PATH)
            # print(f"✅ 已自动保存当前进度至 {SAVE_PATH}")

            # 控制请求频率。
            #time.sleep(4)
            time.sleep(random.uniform(3.2, 6.5))

    if total == 0:
        print("\n接口调用成功，但没有搜索到商品数据。")
        print("请检查关键词、PID、应用权限和多多进宝账号状态。")

    format_worksheet(sheet)

    SAVE_PATH.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    try:
        workbook.save(SAVE_PATH)
    except PermissionError as exc:
        raise RuntimeError(
            f"无法保存文件：{SAVE_PATH}\n"
            "请确认该 Excel 文件没有被其他程序打开。"
        ) from exc

    print("\n" + "=" * 60)
    print(f"抓取结束，共保存 {total} 条商品")
    print(f"Excel 文件：{SAVE_PATH}")
    print("=" * 60)
    # 关闭去重数据库
    duplicate_conn.close()


if __name__ == "__main__":
    try:
        main()

    except Exception as exc:
        print("\n程序执行失败：")
        print(exc)

        print("\n详细错误信息：")
        traceback.print_exc()

        input("\n按回车键关闭窗口……")